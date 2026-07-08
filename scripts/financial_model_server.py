"""
Financial Model API Server
===========================
Deploy this on your VPS alongside financial_model_v5.py.
Exposes a single POST endpoint that n8n (or your frontend) calls.

Usage:
  pip install fastapi uvicorn anthropic yfinance beautifulsoup4 requests openpyxl pandas
  ANTHROPIC_API_KEY="sk-ant-..." python3 financial_model_server.py

The server runs on port 8500 by default.

Deployment setup (one-time, as root):
  sudo mkdir -p /var/lib/financial_models
  sudo chown <deploy-user> /var/lib/financial_models
  # Optional: prevent systemd-tmpfiles from clearing it
  echo "x /var/lib/financial_models" | sudo tee /etc/tmpfiles.d/financial_models.conf
"""

import os
import sys
import uuid
import time
import traceback
from fastapi import FastAPI, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import requests
import uvicorn

# ── Make sure financial_model_v3 is importable from the same directory ──
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from financial_model_v5 import __version__ as FM_VERSION

app = FastAPI(title="Financial Model Generator", version="1.0")

# Output directory for generated models
OUTPUT_DIR = os.environ.get("MODEL_OUTPUT_DIR", "/var/lib/financial_models")
try:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
except PermissionError as e:
    raise RuntimeError(
        f"Cannot create {OUTPUT_DIR}. Run: sudo mkdir -p {OUTPUT_DIR} && sudo chown $(whoami) {OUTPUT_DIR}"
    ) from e
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
MODEL_STORAGE_BUCKET = os.environ.get("SUPABASE_FINANCIAL_MODEL_BUCKET", "research-reports-html")

# Track job status for async mode
jobs: dict[str, dict] = {}


# ========================
# Request / Response Models
# ========================

class GenerateRequest(BaseModel):
    nse_symbol: str
    company_name: str
    sector: str
    folder_id: str | None = None  # Google Drive company vault folder — model xlsx is mirrored here


class GenerateResponse(BaseModel):
    status: str  # "success" | "error"
    file_name: str | None = None
    file_path: str | None = None
    storage_path: str | None = None
    storage_url: str | None = None
    json_storage_path: str | None = None
    json_storage_url: str | None = None
    file_id: str | None = None
    file_url: str | None = None
    message: str | None = None
    duration_seconds: int | None = None


class JobStatus(BaseModel):
    job_id: str
    status: str  # "processing" | "completed" | "failed"
    state: str | None = None  # alias for status, for frontend compatibility
    file_name: str | None = None
    file_path: str | None = None
    storage_path: str | None = None
    storage_url: str | None = None
    json_storage_path: str | None = None
    json_storage_url: str | None = None
    file_id: str | None = None
    file_url: str | None = None
    message: str | None = None
    duration_seconds: int | None = None


class StorageMirrorResponse(BaseModel):
    status: str  # "success" | "error"
    file_name: str | None = None
    file_path: str | None = None
    storage_path: str | None = None
    storage_url: str | None = None
    json_storage_path: str | None = None
    json_storage_url: str | None = None
    message: str | None = None


# ========================
# Health Check
# ========================

@app.get("/health")
def health():
    return {
        "status": "ok",
        "model_version": FM_VERSION,
        "anthropic_key_set": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "screener_creds_set": bool(os.environ.get("SCREENER_USERNAME") and os.environ.get("SCREENER_PASSWORD")),
        "output_dir": OUTPUT_DIR,
    }


def _storage_public_url(path: str) -> str:
    return f"{SUPABASE_URL}/storage/v1/object/public/{MODEL_STORAGE_BUCKET}/{path}"


def upload_model_to_supabase(file_path: str, ticker: str) -> dict[str, str | None]:
    """Upload the v5 xlsx and (when present) the .json sidecar produced by
    financial_model_v5.generate_financial_model. The html-report pipeline
    prefers the JSON sidecar because the xlsx is formula-only and openpyxl
    can't read computed cell values without Excel having recalculated them.
    """
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_KEY not configured")

    ticker_u = ticker.upper()
    path = f"financial-models/{ticker_u}/{ticker_u}_model.xlsx"
    url = f"{SUPABASE_URL}/storage/v1/object/{MODEL_STORAGE_BUCKET}/{path}"
    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "apikey": SUPABASE_SERVICE_KEY,
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "x-upsert": "true",
        "Cache-Control": "no-cache",
    }

    with open(file_path, "rb") as f:
        resp = requests.put(url, data=f, headers=headers, timeout=120)

    if resp.status_code >= 400:
        raise RuntimeError(f"Supabase storage upload failed: {resp.status_code} {resp.text[:400]}")

    # Also upload the JSON sidecar if it exists alongside the xlsx
    json_local = os.path.splitext(file_path)[0] + ".json"
    json_path = None
    json_public_url = None
    if os.path.exists(json_local):
        json_path = f"financial-models/{ticker_u}/{ticker_u}_model.json"
        json_url = f"{SUPABASE_URL}/storage/v1/object/{MODEL_STORAGE_BUCKET}/{json_path}"
        json_headers = {
            **headers,
            "Content-Type": "application/json; charset=utf-8",
        }
        try:
            with open(json_local, "rb") as f:
                jr = requests.put(json_url, data=f, headers=json_headers, timeout=60)
            if jr.status_code >= 400:
                raise RuntimeError(f"JSON sidecar upload failed: {jr.status_code} {jr.text[:200]}")
            print(f"  Uploaded JSON sidecar -> {json_path}")
            json_public_url = _storage_public_url(json_path)
        except Exception as e:  # noqa: BLE001
            raise RuntimeError(f"JSON sidecar upload failed ({type(e).__name__}: {e})") from e
    else:
        print(f"  NOTE: no JSON sidecar at {json_local}; html pipeline will fall back to CSV/xlsx")

    return {
        "storage_path": path,
        "storage_url": _storage_public_url(path),
        "json_storage_path": json_path,
        "json_storage_url": json_public_url,
    }


def upload_model_to_gdrive(file_path: str, ticker: str, folder_id: str | None) -> dict[str, str | None]:
    """Mirror the generated xlsx into the company's Google Drive vault folder
    (same n8n /upload-document webhook used by the PPTX pipeline), into a
    "Financial Model" subfolder. Falls back to uploading straight into the
    company folder if the subfolder-aware call fails. Never raises — a Drive
    upload failure must not fail the overall generation, since the model is
    already safely in Supabase Storage by the time this runs.
    """
    if not folder_id:
        return {"file_id": None, "file_url": None}

    import base64

    n8n_base = os.environ.get("N8N_BASE_URL") or "https://n8n.tikonacapital.com/webhook"
    url = f"{n8n_base.rstrip('/')}/upload-document"
    file_name = f"{ticker.upper()}_model.xlsx"

    with open(file_path, "rb") as fh:
        file_base64 = base64.b64encode(fh.read()).decode("utf-8")

    def _post(subfolder_name: str | None) -> dict[str, str | None] | None:
        payload = {"folder_id": folder_id, "file_name": file_name, "file_base64": file_base64}
        if subfolder_name:
            payload["subfolder_name"] = subfolder_name
        try:
            res = requests.post(url, json=payload, timeout=60)
        except Exception as e:  # noqa: BLE001
            print(f"  Drive upload error (subfolder={subfolder_name}): {e}")
            return None
        if res.status_code != 200:
            print(f"  Drive upload failed (subfolder={subfolder_name}): {res.status_code} {res.text[:200]}")
            return None
        data = res.json()
        if data.get("status") != "success" and "file" not in data and "id" not in data:
            print(f"  Drive upload unexpected response (subfolder={subfolder_name}): {data}")
            return None
        file_info = data.get("file", data)
        if isinstance(file_info, list) and file_info:
            file_info = file_info[0]
        file_id = file_info.get("id")
        file_url = file_info.get("webViewLink") or (
            f"https://drive.google.com/file/d/{file_id}/view" if file_id else None
        )
        return {"file_id": file_id, "file_url": file_url}

    result = _post("Financial Model")
    if result is None:
        print("  Retrying Drive upload directly into company folder (no subfolder)...")
        result = _post(None)

    return result or {"file_id": None, "file_url": None}


# ========================
# Synchronous Generation (n8n calls this with long timeout)
# ========================

@app.post("/generate", response_model=GenerateResponse)
def generate_sync(req: GenerateRequest):
    """
    Synchronous endpoint — blocks until the model is generated.
    n8n should call this with a 15-minute timeout.
    Returns the file path so n8n can read and upload it.
    """
    ticker = req.nse_symbol.strip().upper()
    start = time.time()

    print(f"[generate] Received: nse_symbol='{req.nse_symbol}', ticker='{ticker}', company='{req.company_name}', sector='{req.sector}'")

    if not ticker:
        return GenerateResponse(
            status="error",
            message="nse_symbol is empty — check the webhook payload",
            duration_seconds=0,
        )

    try:
        # v5: formula-based, year-agnostic, cost-tracked
        from financial_model_v5 import generate_financial_model

        out_dir = os.path.join(OUTPUT_DIR, ticker)
        os.makedirs(out_dir, exist_ok=True)

        result = generate_financial_model(
            nse_code=ticker,
            company_name=req.company_name,
            sector=req.sector,
            output_dir=out_dir,
        )

        file_path = result["file_path"]
        json_src = result.get("json_path")

        # Copy xlsx + JSON sidecar into the canonical OUTPUT_DIR slot so
        # upload_model_to_supabase can find both side-by-side.
        final_path = os.path.join(OUTPUT_DIR, f"{ticker}_model.xlsx")
        final_json = os.path.join(OUTPUT_DIR, f"{ticker}_model.json")
        if file_path != final_path:
            import shutil
            shutil.copy2(file_path, final_path)
            file_path = final_path
        if json_src and os.path.exists(json_src) and json_src != final_json:
            import shutil
            shutil.copy2(json_src, final_json)

        storage_result = upload_model_to_supabase(file_path, ticker)
        gdrive_result = upload_model_to_gdrive(file_path, ticker, req.folder_id)

        return GenerateResponse(
            status="success",
            file_name=f"{ticker}_model.xlsx",
            file_path=file_path,
            storage_path=storage_result["storage_path"],
            storage_url=storage_result["storage_url"],
            json_storage_path=storage_result["json_storage_path"],
            json_storage_url=storage_result["json_storage_url"],
            file_id=gdrive_result["file_id"],
            file_url=gdrive_result["file_url"],
            duration_seconds=int(time.time() - start),
            message=(
                f"rating={result.get('rating')} target={result.get('target_price')} "
                f"upside={result.get('upside_pct')}% cost=${result['cost_summary']['cost_usd']}"
            ),
        )

    except Exception as e:
        traceback.print_exc()
        return GenerateResponse(
            status="error",
            message=str(e),
            duration_seconds=int(time.time() - start),
        )


# ========================
# Async Generation (if you want non-blocking)
# ========================

@app.post("/generate-async")
def generate_async(req: GenerateRequest, background_tasks: BackgroundTasks):
    """
    Async endpoint — returns immediately with a job_id.
    Poll /job/{job_id} to check status.
    """
    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {"status": "processing", "started_at": time.time()}

    background_tasks.add_task(_run_generation, job_id, req)

    return {"job_id": job_id, "status": "processing", "message": "Generation started"}


@app.get("/job/{job_id}", response_model=JobStatus)
@app.get("/status/{job_id}", response_model=JobStatus)
def get_job_status(job_id: str):
    """Check the status of an async generation job."""
    if job_id not in jobs:
        return JSONResponse(status_code=404, content={"error": "Job not found"})

    job = jobs[job_id]
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        state=job["status"],
        file_name=job.get("file_name"),
        file_path=job.get("file_path"),
        storage_path=job.get("storage_path"),
        storage_url=job.get("storage_url"),
        json_storage_path=job.get("json_storage_path"),
        json_storage_url=job.get("json_storage_url"),
        message=job.get("message"),
        duration_seconds=job.get("duration_seconds"),
    )


def _run_generation(job_id: str, req: GenerateRequest):
    """Background task for async generation."""
    ticker = req.nse_symbol.upper()
    start = time.time()

    try:
        from financial_model_v5 import generate_financial_model

        result = generate_financial_model(
            nse_code=ticker,
            company_name=req.company_name,
            sector=req.sector,
            output_dir=os.path.join(OUTPUT_DIR, ticker),
        )
        file_path = result["file_path"]
        json_src = result.get("json_path")

        final_path = os.path.join(OUTPUT_DIR, f"{ticker}_model.xlsx")
        final_json = os.path.join(OUTPUT_DIR, f"{ticker}_model.json")
        if file_path != final_path:
            import shutil
            shutil.copy2(file_path, final_path)
            file_path = final_path
        if json_src and os.path.exists(json_src) and json_src != final_json:
            import shutil
            shutil.copy2(json_src, final_json)

        storage_result = upload_model_to_supabase(file_path, ticker)
        gdrive_result = upload_model_to_gdrive(file_path, ticker, req.folder_id)

        jobs[job_id] = {
            "status": "completed",
            "file_name": f"{ticker}_model.xlsx",
            "file_path": file_path,
            "storage_path": storage_result["storage_path"],
            "storage_url": storage_result["storage_url"],
            "json_storage_path": storage_result["json_storage_path"],
            "json_storage_url": storage_result["json_storage_url"],
            "file_id": gdrive_result["file_id"],
            "file_url": gdrive_result["file_url"],
            "duration_seconds": int(time.time() - start),
            "rating": result.get("rating"),
            "target_price": result.get("target_price"),
            "upside_pct": result.get("upside_pct"),
            "cost_usd": result["cost_summary"]["cost_usd"],
        }

    except Exception as e:
        traceback.print_exc()
        jobs[job_id] = {
            "status": "failed",
            "message": str(e),
            "duration_seconds": int(time.time() - start),
        }


# ========================
# Download the generated file
# ========================

@app.get("/download/{ticker}")
def download_model(ticker: str):
    """Download the generated Excel file."""
    ticker = ticker.upper()
    file_path = os.path.join(OUTPUT_DIR, f"{ticker}_model.xlsx")

    if not os.path.exists(file_path):
        return JSONResponse(status_code=404, content={"error": f"No model found for {ticker}"})

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{ticker}_model.xlsx",
    )


@app.post("/storage/{ticker}", response_model=StorageMirrorResponse)
def mirror_model_to_storage(ticker: str):
    ticker = ticker.upper()
    file_path = os.path.join(OUTPUT_DIR, f"{ticker}_model.xlsx")

    if not os.path.exists(file_path):
        return StorageMirrorResponse(
            status="error",
            message=f"No model found for {ticker}",
        )

    try:
        storage_result = upload_model_to_supabase(file_path, ticker)
        return StorageMirrorResponse(
            status="success",
            file_name=f"{ticker}_model.xlsx",
            file_path=file_path,
            storage_path=storage_result["storage_path"],
            storage_url=storage_result["storage_url"],
            json_storage_path=storage_result["json_storage_path"],
            json_storage_url=storage_result["json_storage_url"],
        )
    except Exception as e:
        traceback.print_exc()
        return StorageMirrorResponse(
            status="error",
            file_name=f"{ticker}_model.xlsx",
            file_path=file_path,
            message=str(e),
        )


@app.post("/sync-screener/{ticker}")
def sync_screener(ticker: str):
    """
    Programmatically logs into Screener, downloads the audited financials Excel workbook,
    scrapes live stock indicators from yfinance / Screener HTML, and updates the equity_universe
    database table in Supabase.
    """
    ticker = ticker.strip().upper()
    print(f"[sync-screener] Syncing {ticker} from Screener.in...")
    try:
        username = os.environ.get("SCREENER_USERNAME")
        password = os.environ.get("SCREENER_PASSWORD")
        if not username or not password:
            return {"status": "error", "message": "SCREENER_USERNAME or SCREENER_PASSWORD is not set in server environment"}

        # 1. Download screener excel workbook
        out_dir = os.path.join(OUTPUT_DIR, ticker)
        os.makedirs(out_dir, exist_ok=True)
        
        from financial_model_v5 import download_screener_excel, extract_screener_data
        
        excel_path = download_screener_excel(ticker, out_dir, username, password)
        screener_data = extract_screener_data(excel_path)
        
        # 2. Scrape live market ratios using yfinance / screener HTML
        live_data = {}
        try:
            import yfinance as yf
            info = yf.Ticker(f"{ticker}.NS").info
            live_data["current_price"] = info.get("currentPrice")
            live_data["market_cap"] = info.get("marketCap")
            live_data["pe_ttm"] = info.get("trailingPE")
            live_data["roe"] = (info.get("returnOnEquity") or 0) * 100 if info.get("returnOnEquity") else None
            live_data["fifty_two_week_high"] = info.get("fiftyTwoWeekHigh")
            live_data["fifty_two_week_low"] = info.get("fiftyTwoWeekLow")
            live_data["volume"] = info.get("volume")
        except Exception as e:
            print(f"[sync-screener] yfinance lookup failed for {ticker}: {e}")

        # Supplement with details from Screener HTML page
        from bs4 import BeautifulSoup
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            for u in [
                f"https://www.screener.in/company/{ticker}/consolidated/",
                f"https://www.screener.in/company/{ticker}/",
            ]:
                r = requests.get(u, headers=headers, timeout=10)
                if r.status_code == 200:
                    soup = BeautifulSoup(r.text, "html.parser")
                    top_stats = soup.find("div", class_="company-ratios")
                    if top_stats:
                        for li in top_stats.find_all("li"):
                            name_el = li.find("span", class_="name")
                            val_el = li.find("span", class_="number")
                            if name_el and val_el:
                                name_text = name_el.get_text(strip=True).lower()
                                val_text = val_el.get_text(strip=True).replace(",", "").replace("%", "")
                                try:
                                    val_float = float(val_text)
                                    if "price" in name_text or "current price" in name_text:
                                        if not live_data.get("current_price"):
                                            live_data["current_price"] = val_float
                                    elif "market cap" in name_text:
                                        if not live_data.get("market_cap"):
                                            live_data["market_cap"] = val_float * 1e7  # Screener stores in Cr
                                    elif "stock p/e" in name_text or "p/e" in name_text:
                                        if not live_data.get("pe_ttm"):
                                            live_data["pe_ttm"] = val_float
                                    elif "roce" in name_text:
                                        live_data["roce"] = val_float
                                    elif "roe" in name_text:
                                        if not live_data.get("roe"):
                                            live_data["roe"] = val_float
                                    elif "promoter holding" in name_text:
                                        live_data["promoter_holding_pct"] = val_float
                                    elif "dividend yield" in name_text:
                                        live_data["dividend_yield"] = val_float
                                    elif "face value" in name_text:
                                        live_data["face_value"] = val_float
                                except ValueError:
                                    pass
                    break
        except Exception as e:
            print(f"[sync-screener] Screener HTML parse failed for {ticker}: {e}")

        # 3. Map parsed Excel data to equity_universe columns
        pl = screener_data.get("pl", {})
        bs = screener_data.get("bs", {})
        cf = screener_data.get("cf", {})
        derived = screener_data.get("derived", {})
        fiscal_years = screener_data.get("fiscal_years", [])
        
        def get_fy_val(series, fy):
            try:
                idx = fiscal_years.index(fy)
                v = series[idx]
                return float(v) if v is not None else None
            except (ValueError, IndexError, TypeError):
                return None

        def get_last_val(series):
            if not series:
                return None
            for v in reversed(series):
                if v is not None:
                    return float(v)
            return None

        # Get ISIN from master_company table using nse_symbol (ticker)
        isin_code = None
        if SUPABASE_URL and SUPABASE_SERVICE_KEY:
            headers_api = {
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "apikey": SUPABASE_SERVICE_KEY,
                "Content-Type": "application/json",
            }
            try:
                mc_url = f"{SUPABASE_URL}/rest/v1/master_company?nse_symbol=eq.{ticker}&select=isin"
                mc_resp = requests.get(mc_url, headers=headers_api, timeout=10)
                if mc_resp.status_code == 200 and mc_resp.json():
                    isin_code = mc_resp.json()[0].get("isin")
            except Exception as e:
                print(f"[sync-screener] master_company lookup failed for {ticker}: {e}")
        
        # Fallback to avoid violating not-null constraint on equity_universe.isin_code
        if not isin_code:
            isin_code = f"INE{ticker:<9}".replace(" ", "0")[:12]

        payload = {
            "company_name": screener_data.get("company_name"),
            "nse_code": ticker,
            "isin_code": isin_code,
            "current_price": live_data.get("current_price") or screener_data.get("cmp"),
            "market_cap": live_data.get("market_cap") or screener_data.get("mcap"),
            "pe_ttm": live_data.get("pe_ttm"),
            "roe": live_data.get("roe"),
            "roce": live_data.get("roce"),
            "promoter_holding_pct": live_data.get("promoter_holding_pct"),
            "dividend_yield": live_data.get("dividend_yield"),
            "face_value": live_data.get("face_value") or screener_data.get("face_value"),
            
            # Historical Revenue
            "revenue_fy2023": get_fy_val(pl.get("sales"), "FY23"),
            "revenue_fy2024": get_fy_val(pl.get("sales"), "FY24"),
            "revenue_fy2025": get_fy_val(pl.get("sales"), "FY25"),
            "revenue_ttm": get_last_val(pl.get("sales")),
            
            # Historical PAT
            "pat_fy2023": get_fy_val(pl.get("net_profit"), "FY23"),
            "pat_fy2024": get_fy_val(pl.get("net_profit"), "FY24"),
            "pat_fy2025": get_fy_val(pl.get("net_profit"), "FY25"),
            "pat_ttm": get_last_val(pl.get("net_profit")),
            
            # Balance Sheet / Debt / Cash
            "debt": get_last_val(bs.get("borrowings")),
            "cash_equivalents": get_last_val(bs.get("cash")),
            "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S")
        }
        
        # Calculate derived margins
        sales_ttm = payload["revenue_ttm"]
        pat_ttm = payload["pat_ttm"]
        if sales_ttm and pat_ttm:
            payload["pat_margin_ttm"] = (pat_ttm / sales_ttm) * 100
        
        ebitda_ttm = get_last_val(derived.get("ebitda"))
        if sales_ttm and ebitda_ttm:
            payload["ebitda_margin_ttm"] = (ebitda_ttm / sales_ttm) * 100

        # 4. Write to Supabase using REST API
        if SUPABASE_URL and SUPABASE_SERVICE_KEY:
            headers = {
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "apikey": SUPABASE_SERVICE_KEY,
                "Content-Type": "application/json",
            }
            # Query existing company_id
            query_url = f"{SUPABASE_URL}/rest/v1/equity_universe?nse_code=eq.{ticker}&select=company_id"
            q_resp = requests.get(query_url, headers=headers, timeout=10)
            if q_resp.status_code == 200 and q_resp.json():
                company_id = q_resp.json()[0]["company_id"]
                # Update (PATCH)
                up_url = f"{SUPABASE_URL}/rest/v1/equity_universe?company_id=eq.{company_id}"
                up_resp = requests.patch(up_url, json=payload, headers=headers, timeout=10)
                if up_resp.status_code >= 400:
                    raise Exception(f"Failed to update equity_universe: {up_resp.status_code} {up_resp.text}")
                payload["company_id"] = company_id
            else:
                # Insert (POST)
                ins_url = f"{SUPABASE_URL}/rest/v1/equity_universe"
                ins_resp = requests.post(ins_url, json=payload, headers=headers, timeout=10)
                if ins_resp.status_code >= 400:
                    raise Exception(f"Failed to insert into equity_universe: {ins_resp.status_code} {ins_resp.text}")

        return {"status": "success", "data": payload}

    except Exception as e:
        traceback.print_exc()
        return {"status": "error", "message": str(e)}


@app.post("/regenerate-json/{ticker}")
def regenerate_json(ticker: str):
    """
    Recalculates the formulas of the user-modified Excel model on Supabase Storage
    via LibreOffice Calc, then parses and updates the corresponding JSON sidecar,
    fetching live CMP from yfinance, and uploads both back to Supabase Storage.
    """
    ticker = ticker.strip().upper()
    print(f"[/regenerate-json] Received request for ticker: {ticker}")
    if not ticker:
        return JSONResponse(status_code=400, content={"status": "error", "message": "Ticker is empty"})

    import json
    import uuid
    import shutil
    import subprocess
    from pathlib import Path
    import openpyxl
    from openpyxl.utils import get_column_letter
    import yfinance as yf

    temp_id = str(uuid.uuid4())
    temp_dir = os.path.join(OUTPUT_DIR, f"regen_{temp_id}")
    os.makedirs(temp_dir, exist_ok=True)
    excel_path = os.path.join(temp_dir, f"{ticker}_model.xlsx")
    json_path = os.path.join(temp_dir, f"{ticker}_model.json")

    try:
        # 1. Download current files from Supabase
        xlsx_url = f"{SUPABASE_URL}/storage/v1/object/{MODEL_STORAGE_BUCKET}/financial-models/{ticker}/{ticker}_model.xlsx"
        json_url = f"{SUPABASE_URL}/storage/v1/object/{MODEL_STORAGE_BUCKET}/financial-models/{ticker}/{ticker}_model.json"
        
        headers = {
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "apikey": SUPABASE_SERVICE_KEY,
        }
        
        print(f"Downloading Excel from {xlsx_url}...")
        r_xlsx = requests.get(xlsx_url, headers=headers, timeout=60)
        if r_xlsx.status_code >= 400:
            return JSONResponse(status_code=400, content={"status": "error", "message": f"Failed to download Excel from Supabase: {r_xlsx.status_code}"})
        with open(excel_path, "wb") as f:
            f.write(r_xlsx.content)

        print(f"Downloading JSON from {json_url}...")
        r_json = requests.get(json_url, headers=headers, timeout=60)
        model_json = {}
        if r_json.status_code < 400:
            try:
                model_json = json.loads(r_json.content.decode("utf-8"))
            except Exception as e:
                print(f"Error parsing existing JSON: {e}")

        # 2. Run LibreOffice calculation on excel_path
        soffice_bin = shutil.which("soffice") or shutil.which("soffice.exe")
        if not soffice_bin:
            for win_path in [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]:
                if os.path.exists(win_path):
                    soffice_bin = win_path
                    break

        if not soffice_bin:
            print("LibreOffice (soffice) not found; formula recalculation skipped.")
        else:
            excel_path_p = Path(excel_path)
            raw_dir = excel_path_p.parent / "raw_input"
            raw_dir.mkdir(parents=True, exist_ok=True)
            raw_path = raw_dir / excel_path_p.name
            try:
                shutil.move(str(excel_path), str(raw_path))
                cmd = [
                    soffice_bin,
                    "--headless",
                    "--convert-to", "xlsx",
                    "--outdir", str(excel_path_p.parent),
                    str(raw_path)
                ]
                print("Running LibreOffice recalculation:", " ".join(cmd))
                subprocess.run(cmd, capture_output=True, text=True, timeout=60, check=True)
            except Exception as exc:
                print("Failed to recalculate formulas via LibreOffice:", exc)
                if raw_path.exists() and not excel_path_p.exists():
                    shutil.copy2(str(raw_path), str(excel_path_p))
            finally:
                if raw_dir.exists():
                    try:
                        shutil.rmtree(raw_dir)
                    except Exception:
                        pass

        # 3. Load Excel and update model_json fields
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Financial Summary Sheet
        if "Financial Summary" in wb.sheetnames:
            ws_fs = wb["Financial Summary"]
            shares_cr = ws_fs["B9"].value
            if shares_cr is not None:
                model_json["shares_cr"] = float(shares_cr)
            base_year = ws_fs["B10"].value
            if base_year is not None:
                model_json["base_year"] = str(base_year)
            rating = ws_fs["B7"].value
            if rating is not None:
                model_json["rating"] = str(rating)
            target_price = ws_fs["B5"].value
            if target_price is not None:
                model_json["target_price"] = float(target_price)
            upside_val = ws_fs["B6"].value
            if upside_val is not None:
                model_json["upside_pct"] = round(float(upside_val) * 100, 2)
            market_cap = ws_fs["B8"].value
            if market_cap is not None:
                model_json["market_cap"] = float(market_cap)

        # Assumptions Sheet
        if "Assumptions" in wb.sheetnames:
            ws_as = wb["Assumptions"]
            nc = ws_as.max_column
            h_cols = []
            p_cols = []
            for col in range(2, nc + 1):
                val = ws_as.cell(row=9, column=col).value
                if val:
                    val_str = str(val)
                    if val_str.endswith("A"):
                        h_cols.append(col)
                    elif val_str.endswith("E"):
                        p_cols.append(col)

            proj_years = [str(ws_as.cell(row=9, column=col).value) for col in p_cols]
            if "assumptions" not in model_json:
                model_json["assumptions"] = {}
            if proj_years:
                model_json["assumptions"]["projection_years"] = proj_years

            def _safe_float(val_in):
                try:
                    return float(val_in) if val_in not in (None, "", "—") else 0.0
                except Exception:
                    return 0.0

            model_json["assumptions"]["wacc_pct"] = _safe_float(ws_as["B42"].value)
            model_json["assumptions"]["terminal_growth_pct"] = _safe_float(ws_as["B43"].value)
            model_json["assumptions"]["target_pe"] = _safe_float(ws_as["B44"].value)

            target_ev = ws_as["B45"].value
            if target_ev not in (None, "", "—"):
                try:
                    model_json["assumptions"]["target_ev_ebitda"] = float(target_ev)
                except Exception:
                    model_json["assumptions"]["target_ev_ebitda"] = None
            else:
                model_json["assumptions"]["target_ev_ebitda"] = None

            model_json["assumptions"]["dividend_payout_pct"] = _safe_float(ws_as["B46"].value)
            model_json["assumptions"]["rationale"] = ws_as.cell(row=49, column=1).value or ""

            assum_items = [
                ("revenue_growth_pct", 10),
                ("rm_pct", 11),
                ("employee_pct", 12),
                ("power_fuel_pct", 13),
                ("other_mfg_pct", 14),
                ("selling_admin_pct", 15),
                ("other_exp_pct", 16),
                ("chg_inventory_pct", 17),
                ("depreciation_cr", 18),
                ("interest_cr", 19),
                ("other_income_cr", 20),
                ("tax_rate_pct", 21),
                ("capex_cr", 22),
                ("receivable_days", 23),
                ("inventory_days", 24),
                ("other_assets_pct", 25),
            ]
            for key, r_idx in assum_items:
                model_json["assumptions"][key] = {}
                for col, yr in zip(p_cols, proj_years):
                    val = ws_as.cell(row=r_idx, column=col).value
                    if val is not None:
                        try:
                            model_json["assumptions"][key][yr] = float(val)
                        except Exception:
                            pass

            bs_cf_items = [
                ("share_capital", 29),
                ("reserves", 30),
                ("borrowings", 31),
                ("other_liabilities", 32),
                ("net_block", 33),
                ("cwip", 34),
                ("investments", 35),
                ("other_assets", 36),
                ("cfo", 37),
                ("cfi", 38),
                ("cff", 39),
            ]
            if "projections" not in model_json:
                model_json["projections"] = {}
            model_json["projections"]["years"] = proj_years
            for key, r_idx in bs_cf_items:
                vals = []
                for col in p_cols:
                    val = ws_as.cell(row=r_idx, column=col).value
                    try:
                        vals.append(float(val) if val is not None else None)
                    except Exception:
                        vals.append(None)
                model_json["projections"][key] = vals

        # P&L Sheet
        if "P&L" in wb.sheetnames:
            ws_pl = wb["P&L"]
            proj_rev = []
            proj_ebitda = []
            proj_pat = []
            proj_eps = []
            for col in p_cols:
                try:
                    proj_rev.append(float(ws_pl.cell(row=3, column=col).value or 0))
                    proj_ebitda.append(float(ws_pl.cell(row=12, column=col).value or 0))
                    proj_pat.append(float(ws_pl.cell(row=20, column=col).value or 0))
                    proj_eps.append(float(ws_pl.cell(row=23, column=col).value or 0))
                except Exception:
                    proj_rev.append(0.0)
                    proj_ebitda.append(0.0)
                    proj_pat.append(0.0)
                    proj_eps.append(0.0)

            model_json["projected_pl"] = {
                "years": proj_years,
                "revenue": [round(x, 2) for x in proj_rev],
                "ebitda": [round(x, 2) for x in proj_ebitda],
                "pat": [round(x, 2) for x in proj_pat],
                "eps": [round(x, 4) for x in proj_eps],
            }

        # SAARTHI Sheet
        if "SAARTHI" in wb.sheetnames:
            ws_saarthi = wb["SAARTHI"]
            dimensions = []
            expected_keys = ["S", "A1", "A2", "R", "T", "H", "I"]
            for i, key in enumerate(expected_keys, start=5):
                try:
                    dimensions.append({
                        "key": key,
                        "name": ws_saarthi.cell(row=i, column=2).value,
                        "score": float(ws_saarthi.cell(row=i, column=3).value or 0),
                        "max_score": float(ws_saarthi.cell(row=i, column=4).value or 0),
                        "rationale": ws_saarthi.cell(row=i, column=6).value or "",
                    })
                except Exception:
                    pass
            if "saarthi_scorecard" not in model_json:
                model_json["saarthi_scorecard"] = {}
            model_json["saarthi_scorecard"]["dimensions"] = dimensions
            try:
                model_json["saarthi_scorecard"]["total_score"] = int(ws_saarthi.cell(row=12, column=3).value or 70)
            except Exception:
                model_json["saarthi_scorecard"]["total_score"] = 70

        # Scenario Analysis Sheet
        if "Scenario_Analysis" in wb.sheetnames:
            ws_scen = wb["Scenario_Analysis"]
            def _read_scen(col_letter):
                try:
                    adj = ws_scen[f"{col_letter}7"].value
                    pe = ws_scen[f"{col_letter}8"].value
                    prob = ws_scen[f"{col_letter}9"].value
                    tp = ws_scen[f"{col_letter}10"].value
                    return {
                        "eps_adjustment_pct": round(float(adj) * 100, 2) if adj is not None else 0.0,
                        "target_pe": float(pe) if pe is not None else 0.0,
                        "probability_pct": round(float(prob) * 100, 2) if prob is not None else 0.0,
                        "target_price": float(tp) if tp is not None else 0.0,
                    }
                except Exception:
                    return {"eps_adjustment_pct": 0.0, "target_pe": 0.0, "probability_pct": 0.0, "target_price": 0.0}

            try:
                weighted_tp = float(ws_scen["B13"].value or 0.0)
            except Exception:
                weighted_tp = 0.0

            model_json["scenario_analysis"] = {
                "bull": _read_scen("B"),
                "base": _read_scen("C"),
                "bear": _read_scen("D"),
                "weighted_tp": weighted_tp,
            }

        wb.close()

        # 4. Fetch live CMP from yfinance and update cmp/upside_pct
        try:
            info = yf.Ticker(f"{ticker}.NS").info
            live_cmp = info.get("currentPrice")
            if live_cmp:
                model_json["cmp"] = float(live_cmp)
                if model_json.get("target_price"):
                    model_json["upside_pct"] = round((model_json["target_price"] / live_cmp - 1) * 100, 2)
                if model_json.get("shares_cr"):
                    model_json["market_cap"] = round(live_cmp * model_json["shares_cr"], 2)
        except Exception as e:
            print(f"yfinance fetch failed during JSON regeneration: {e}")

        # 5. Write regenerated JSON locally
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(model_json, f, indent=2, default=str)

        # 6. Upload Excel + JSON back to Supabase Storage
        storage_result = upload_model_to_supabase(excel_path, ticker)

        return {
            "status": "success",
            "json_storage_url": storage_result["json_storage_url"],
            "storage_url": storage_result["storage_url"],
        }

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"status": "error", "message": f"JSON regeneration failed: {str(e)}"})
    finally:
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass


# ========================
# Run Server
# ========================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8500))
    print(f"\n🚀 Financial Model Server starting on port {port}")
    print(f"📁 Output directory: {OUTPUT_DIR}")
    print(f"🔑 Anthropic key: {'✓ Set' if os.environ.get('ANTHROPIC_API_KEY') else '✗ MISSING'}\n")

    uvicorn.run(app, host="0.0.0.0", port=port)
